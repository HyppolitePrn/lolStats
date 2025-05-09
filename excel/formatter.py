import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import COLORS

logger = logging.getLogger(__name__)

def apply_header_style(cell):
    """Apply header styling to a cell"""
    cell.fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='center')

def apply_day_style(cell):
    """Apply day header styling to a cell"""
    cell.fill = PatternFill(start_color=COLORS["day"], end_color=COLORS["day"], fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

def apply_match_style(cell):
    """Apply match header styling to a cell"""
    cell.fill = PatternFill(start_color=COLORS["match"], end_color=COLORS["match"], fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

def apply_win_loss_style(cell, is_win):
    """Apply win/loss styling to a cell"""
    if is_win:
        cell.fill = PatternFill(start_color=COLORS["win"], end_color=COLORS["win"], fill_type="solid")
    else:
        cell.fill = PatternFill(start_color=COLORS["loss"], end_color=COLORS["loss"], fill_type="solid")

def adjust_columns_width(worksheet, min_width=10, max_width=40):
    """Auto-adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = max(min_width, min(max_length + 2, max_width))
        worksheet.column_dimensions[column_letter].width = adjusted_width