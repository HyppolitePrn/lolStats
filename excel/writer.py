import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import COLORS, SHEET_NAME, EXCEL_HEADERS
from stats.calculator import sort_players_by_position

logger = logging.getLogger(__name__)

def get_processed_matches(wb):
    """Get a set of already processed match IDs from the Excel workbook"""
    processed_matches = set()
    
    try:
        # Try to get previously processed matches from a hidden sheet
        if "ProcessedMatches" in wb.sheetnames:
            ws = wb["ProcessedMatches"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Check if match ID exists
                    processed_matches.add(row[0])
        
        logger.info(f"Found {len(processed_matches)} already processed match IDs")
        return processed_matches
    
    except Exception as e:
        logger.error(f"Error getting processed matches: {str(e)}")
        return set()

def add_processed_match(wb, match_id):
    """Add a match ID to the processed matches sheet"""
    try:
        # Create or get the ProcessedMatches sheet
        if "ProcessedMatches" not in wb.sheetnames:
            ws = wb.create_sheet("ProcessedMatches")
            ws.sheet_state = 'hidden'  # Hide the sheet
            ws['A1'] = "MatchID"
        else:
            ws = wb["ProcessedMatches"]
        
        # Add the match ID
        ws.append([match_id])
        logger.debug(f"Added match ID to processed list: {match_id}")
    
    except Exception as e:
        logger.error(f"Error adding processed match: {str(e)}")

def update_excel_with_stats(excel_path, all_match_stats):
    """Update an existing Excel file or create a new one with team stats"""
    logger.info(f"Updating Excel file: {excel_path}")
    logger.info(f"Number of match stats to add: {len(all_match_stats)}")
    
    # Define styles
    header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    day_fill = PatternFill(start_color=COLORS["day"], end_color=COLORS["day"], fill_type="solid")
    match_fill = PatternFill(start_color=COLORS["match"], end_color=COLORS["match"], fill_type="solid")
    win_fill = PatternFill(start_color=COLORS["win"], end_color=COLORS["win"], fill_type="solid")
    loss_fill = PatternFill(start_color=COLORS["loss"], end_color=COLORS["loss"], fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Define column headers
    headers = EXCEL_HEADERS
    
    # Check if the file exists
    file_exists = Path(excel_path).exists()
    
    try:
        if file_exists:
            # Load existing workbook
            logger.info(f"Loading existing Excel file: {excel_path}")
            wb = load_workbook(excel_path)
            
            # Get active sheet or create one
            if wb.active:
                ws = wb.active
            else:
                ws = wb.create_sheet(SHEET_NAME)
                logger.info(f"Created new sheet: {SHEET_NAME}")
            
            # Get the set of matches we've already processed
            processed_matches = get_processed_matches(wb)
        else:
            # Create a new workbook
            logger.info(f"Creating new Excel file: {excel_path}")
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            processed_matches = set()
            
            # Initialize headers if it's a new file
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # Find the last row with data
        last_row = 1
        for row in ws.iter_rows():
            if all(cell.value is None for cell in row):
                break
            last_row += 1
        
        logger.info(f"Last row in Excel: {last_row}")
        
        # Group new data by day and match
        days = {}
        for match_stats in all_match_stats:
            # Skip already processed matches
            match_id = match_stats["match_id"]
            if match_id in processed_matches:
                logger.info(f"Skipping already processed match: {match_id}")
                continue
            
            day = match_stats["day"]
            match_num = match_stats["match"]
            team_stats = match_stats["team_stats"]
            
            # Sort players by position
            team_stats = sort_players_by_position(team_stats)
            
            logger.debug(f"Adding day: {day}, match: {match_num}, team stats: {len(team_stats)}")
            
            if day not in days:
                days[day] = {}
            
            if match_num not in days[day]:
                days[day][match_num] = []
            
            days[day][match_num].append(team_stats)
            
            # Mark this match as processed
            add_processed_match(wb, match_id)
        
        logger.info(f"Days to process: {len(days)}")
        
        if not days:
            logger.warning("No new data to add to Excel file")
            logger.info(f"Saving workbook to {excel_path}")
            wb.save(excel_path)
            return
        
        # Start adding new data from the last row
        row = last_row
        
        # Iterate through days and matches
        for day, matches in sorted(days.items()):
            logger.debug(f"Processing day {day} with {len(matches)} matches")
            
            # Write day header
            # Write day header
            day_label = f"Day {day}"
            ws.cell(row=row, column=1, value=day_label).fill = day_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            for match_num, match_data in sorted(matches.items()):
                logger.debug(f"Processing match {match_num} with {len(match_data)} teams")
                
                # Write match header
                match_label = f"Match {match_num}"
                ws.cell(row=row, column=1, value=match_label).fill = match_fill
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                # Write column headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                row += 1
                
                # Write player data for this match
                for team_stats in match_data:
                    logger.debug(f"Processing team with {len(team_stats)} players")
                    
                    for player in team_stats:
                        # Get win/loss status for cell color
                        is_win = player["win"]
                        fill = win_fill if is_win else loss_fill
                        
                        # Write player stats
                        ws.cell(row=row, column=1, value=player["summonerName"])
                        ws.cell(row=row, column=2, value=player["champion"])
                        ws.cell(row=row, column=3, value=player["position"])
                        ws.cell(row=row, column=4, value=f"{player['kills']}/{player['deaths']}/{player['assists']}")
                        ws.cell(row=row, column=5, value=player["kda"])
                        ws.cell(row=row, column=6, value=player["DPM"])
                        ws.cell(row=row, column=7, value=player["VPM"])
                        ws.cell(row=row, column=8, value=player["CSperMin"])
                        ws.cell(row=row, column=9, value=player["goldDiffAt15"])
                        ws.cell(row=row, column=10, value=player["expDiffAt15"])
                        ws.cell(row=row, column=11, value=player["soloKills"])
                        ws.cell(row=row, column=12, value=player["killParticipation"])
                        ws.cell(row=row, column=13, value="Win" if is_win else "Loss").fill = fill
                        
                        # Add border to all cells
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=row, column=col).border = thin_border
                        
                        row += 1
                    
                    # Add an empty row between teams
                    row += 1
                
                # Add an empty row between matches
                row += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = 15
        
        # Save the workbook
        logger.info(f"Saving workbook to {excel_path}")
        wb.save(excel_path)
        logger.info(f"Excel file updated: {excel_path}")
    
    except Exception as e:
        logger.error(f"Error updating Excel: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())